# bot.py
import os
import re
import sys
import io
import json
import hashlib
import time
import asyncio
import threading
import subprocess
import shutil
import uuid
import pandas as pd
import unicodedata
from http.server import BaseHTTPRequestHandler, HTTPServer
from urllib.parse import urlparse
from telegram import (
    Update,
    ReplyKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardRemove,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    CallbackQueryHandler,
    filters,
)
from telegram.error import NetworkError, TimedOut
from PyPDF2 import PdfReader, PdfWriter

# ضمان طباعة عربية مباشرة
try:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
except Exception:
    pass



# ==========================
# 🧹 تطبيع خفيف للبحث داخل PDF (مطابق لاختبارك الذي نجح)
# ==========================
def norm_phrase(s: str) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = s.translate(str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789"))
    s = re.sub(r"[\u200f\u200e\u200b\u202a\u202b\u202c\u202d\u202e]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s
# ==========================
# 🧹 دالة تطبيع النص العربي (V2)
# ==========================
def normalize_arabic_text_v2(s: str) -> str:
    if not s:
        return ""

    # 1) تحويل Presentation Forms → Unicode قياسي
    s = unicodedata.normalize("NFKC", s)

    # 2) تحويل الأرقام العربية الهندية → إنجليزية
    s = s.translate(str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789"))

    # 3) إزالة المحارف الخفية بالكامل
    s = re.sub(r"[\u200f\u200e\u200b\u202a\u202b\u202c\u202d\u202e]", "", s)

    # 4) إزالة التطويل والتشكيل
    s = re.sub(r"[ـًٌٍَُِّْٰ]", "", s)

    # 5) إزالة كل شيء غير رقمي بين أرقام داخل نفس الرقم
    # مثال: 11‏7‏595‏9584 → 1175959584
    s = re.sub(r"(?<=\d)[^\d]+(?=\d)", "", s)

    return s.strip()

# =========================
# إعدادات أساسية
# =========================
FILES = {
    "schedule": "Scheduals.pdf",
    "advisor": "Advisors.csv",
    "remaining": "Remaining.pdf",
    "gpa": "GPA.pdf",
    "majors": "TNumbers with majors.pdf",
    "ids": "IDs.csv",
    "certificates": "Certificates.pdf",
    "permission": "permission_to_conduct_research.pdf",
    "aramco_training": os.path.join(
        "خطابات_طلب_فرصة_تدريبية_في_شركة_ارامكو_للفصل_الأول_1448.pdf",
        "خطابات_طلب_فرصة_تدريبية_في_شركة_ارامكو_للفصل_الأول_1448.pdf",
    ),
}


# ✅ استخدم متغير البيئة TELEGRAM_TOKEN
BOT_TOKEN = os.environ.get("TELEGRAM_TOKEN")
if not BOT_TOKEN:
    print("❌ لم يتم العثور على متغير TELEGRAM_TOKEN. ضعه في إعدادات الخادم أو عرّفه محليًا للتجربة.", flush=True)
    sys.exit(1)

# =========================
# حالة البوت (تُعرض للداثبورد)
# =========================
STATUS = {
    "running": False,
    "telegram_connected": False,
    "indexes_ready": False,
    "indexing": False,
    "current_file": "",
    "index_progress": 0.0,   # 0..100
    "last_user": ""
}
_status_lock = threading.Lock()
INDEXES_READY = threading.Event()
SESSION_TTL_SECONDS = 30 * 60
TEMP_ROOT = os.environ.get("BOT_TEMP_DIR", "tmp")

def _set_status(**kwargs):
    with _status_lock:
        STATUS.update(kwargs)

def _mask_identifier(value: str, visible_prefix: int = 3, visible_suffix: int = 3) -> str:
    value = str(value or "")
    if len(value) <= visible_prefix + visible_suffix:
        return "*" * len(value)
    return f"{value[:visible_prefix]}***{value[-visible_suffix:]}"

def _log_user_input(txt: str, context: ContextTypes.DEFAULT_TYPE):
    normalized = convert_arabic_to_english(txt)
    if re.fullmatch(r"44\d{7}", normalized):
        event = f"student_id={_mask_identifier(normalized)}"
    elif re.fullmatch(r"[0-9٠-٩]{10}", txt):
        event = "national_id=<hidden>"
    elif txt.startswith("/"):
        event = f"command={txt.split(maxsplit=1)[0]}"
    elif txt:
        event = f"text_length={len(txt)}"
    else:
        event = "empty_message"

    _set_status(last_user=event)
    print(f"💬 المستخدم: {event}", flush=True)

async def _reply_indexes_not_ready(update: Update):
    status = _get_status()
    current_file = status.get("current_file") or "ملفات البيانات"
    progress = status.get("index_progress") or 0.0
    await update.message.reply_text(
        "⏳ البوت يجهز بيانات الطلاب الآن. حاول بعد قليل.\n"
        f"الملف الحالي: {current_file} ({progress:.0f}%)"
    )

async def _ensure_indexes_ready(update: Update) -> bool:
    if INDEXES_READY.is_set():
        return True
    await _reply_indexes_not_ready(update)
    return False

def _mark_authenticated(context: ContextTypes.DEFAULT_TYPE, student_id: str):
    context.user_data["student_id"] = student_id
    context.user_data["authenticated_at"] = time.time()

async def _expire_session_if_needed(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    if "student_id" not in context.user_data:
        return False

    authenticated_at = context.user_data.get("authenticated_at")
    if not authenticated_at:
        context.user_data["authenticated_at"] = time.time()
        return False

    if time.time() - float(authenticated_at) <= SESSION_TTL_SECONDS:
        return False

    context.user_data.clear()
    await update.message.reply_text(
        "⏳ انتهت جلسة الدخول. أرسل رقمك التدريبي لتسجيل الدخول من جديد.",
        reply_markup=ReplyKeyboardRemove()
    )
    return True

_fingerprint_cache = {}

def _file_fingerprint(path: str):
    stat = os.stat(path)
    abs_path = os.path.abspath(path)
    cache_key = (abs_path, stat.st_size, stat.st_mtime_ns)
    cached = _fingerprint_cache.get(cache_key)
    if cached:
        return cached

    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)

    fingerprint = {
        "size": stat.st_size,
        "sha256": h.hexdigest(),
    }
    _fingerprint_cache[cache_key] = fingerprint
    return fingerprint

def _load_index_if_current(pdf_path: str, index_path: str, ready_message: str):
    meta_path = index_path + ".meta"
    if not (os.path.exists(pdf_path) and os.path.exists(index_path) and os.path.exists(meta_path)):
        return None

    try:
        with open(meta_path, "r", encoding="utf-8") as m:
            meta = json.load(m)
    except Exception:
        return None

    if not isinstance(meta, dict):
        return None

    fingerprint = _file_fingerprint(pdf_path)
    if meta.get("size") == fingerprint["size"] and meta.get("sha256") == fingerprint["sha256"]:
        print(ready_message, flush=True)
        with open(index_path, "r", encoding="utf-8") as f:
            return json.load(f)

    return None

def _write_index_meta(pdf_path: str, index_path: str):
    meta_path = index_path + ".meta"
    with open(meta_path, "w", encoding="utf-8") as m:
        json.dump(_file_fingerprint(pdf_path), m, ensure_ascii=False)


def build_majors_plan_index(pdf_path, index_path="majors_plan_index.json"):
    """
    ✅ فهرس سريع للخطة التفصيلية:
        44xxxxxxx -> اسم ملف الخطة (HRplan/EPplan/...)

    - يستخدم norm_phrase (تطبيع خفيف) لأنه أثبت نجاحه في مطابقة العبارات داخل PDF.
    - يربط IDs الموجودة في نفس الصفحة التي تحتوي عبارة التخصص.
    - يستخدم .meta حتى لا يعيد البناء إلا إذا تغيّر الـ PDF.
    """
    try:
        cached = _load_index_if_current(pdf_path, index_path, "✅ فهرس خطط التخصصات جاهز مسبقًا.")
        if cached is not None:
            return cached

        if not os.path.exists(pdf_path):
            print(f"❌ ملف التخصصات غير موجود: {pdf_path}", flush=True)
            return {}

        print(f"⏳ بناء فهرس خطط التخصصات (نفس الصفحة) من: {pdf_path}", flush=True)

        reader = PdfReader(pdf_path)
        total_pages = len(reader.pages)

        sid_re = re.compile(r"\b44\d{7}\b")

        out = {}

        # تجهيز عبارات مطبّعة مسبقًا
        norm_phrases = [(norm_phrase(k), v) for k, v in MAJOR_PHRASES_TO_PLAN.items()]

        for i, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            nt = norm_phrase(text)

            # تحديد الخطة من نفس الصفحة
            plan = None
            for ph, plan_file in norm_phrases:
                if ph and ph in nt:
                    plan = plan_file
                    break

            # ربط IDs بالخطة
            if plan:
                ids = set(sid_re.findall(nt))
                for sid in ids:
                    out[sid] = plan

            if (i + 1) % 50 == 0 or (i + 1) == total_pages:
                percent = (i + 1) / max(1, total_pages) * 100
                print(f"📄 MAJORS PLAN: {i+1}/{total_pages} ({percent:.1f}%) | mapped_ids={len(out)}", flush=True)

        with open(index_path, "w", encoding="utf-8") as f:
            json.dump(out, f, ensure_ascii=False)

        _write_index_meta(pdf_path, index_path)

        print(f"✅ تم بناء فهرس خطط التخصصات: {len(out)} متدرب.", flush=True)
        return out

    except Exception as e:
        print("❌ خطأ أثناء بناء فهرس خطط التخصصات:", e, flush=True)
        import traceback
        traceback.print_exc()
        return {}




def _get_status():
    with _status_lock:
        return dict(STATUS)

# =========================
# أدوات مساعدة
# =========================
def convert_arabic_to_english(arabic_number: str) -> str:
    arabic_digits = {
        '٠': '0','١': '1','٢': '2','٣': '3','٤': '4',
        '٥': '5','٦': '6','٧': '7','٨': '8','٩': '9'
    }
    return ''.join(arabic_digits.get(ch, ch) for ch in arabic_number)

# =============== أدوات مساعدة للهوية والاسم ===============
AR_DIGITS = str.maketrans('٠١٢٣٤٥٦٧٨٩', '0123456789')

def normalize_digits(s: str) -> str:
    """تحويل الأرقام العربية-الهندية إلى إنجليزية + إزالة محارف خفية."""
    return (s or "").translate(AR_DIGITS).replace('\u200f', '').replace('\u200e', '').strip()

def is_valid_nid(nid: str) -> bool:
    """هوية وطنية سعودية: تبدأ بـ 1 وطولها 10 أرقام."""
    nid = normalize_digits(nid)
    return bool(re.fullmatch(r'1\d{9}', nid))

def looks_like_ar_name(line: str) -> bool:
    if not re.search(r'[اأإآء-ي]', line):
        return False
    s = re.sub(r'[^اأإآء-ي\s]', '', line).strip()
    if not s:
        return False
    words = [w for w in s.split() if len(w) >= 2]
    return 2 <= len(words) <= 5

def clean_ar_name(line: str) -> str:
    s = re.sub(r'[^\w\s\u0600-\u06FF]', ' ', line)  # إبقاء العربية والمسافات
    s = re.sub(r'\s+', ' ', s).strip()
    # إزالة ألفاظ وصفية شائعة لو ظهرت ملتصقة بالاسم
    s = re.sub(r'\b(الطالب|المتدرب|اسم|Name)\b', '', s).strip()
    return s

def extract_first_name(full_name: str) -> str:
    """
    أول اسم منطقي مع مراعاة المركّب (عبد الرحمن، عبد الله).
    لو ما قدر، يرجع أول كلمة سليمة.
    """
    full_name = clean_ar_name(full_name)
    parts = full_name.split()
    if not parts:
        return "عزيزنا"
    if len(parts) >= 2 and parts[0] in ("عبد", "أبو", "أم", "ابن", "بن"):
        return f"{parts[0]} {parts[1]}"
    return parts[0]

# =========================
# 🟦 أدوات حساب الأسبوع الحالي
# =========================
import openpyxl
from datetime import datetime

EXCEL_PATH = "Calender.xlsx"   # تأكد أنه بجانب bot.py

def load_weeks_from_excel(sheet_name):
    """تحميل الأسابيع من شيت محدد (Term1 أو Term2)."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[sheet_name]

    weeks = {}

    for col in range(1, ws.max_column + 1):
        week_dates = []
        for row in range(2, 7):  # 5 أيام الأسبوع
            value = ws.cell(row=row, column=col).value
            if value and not value.startswith("—"):
                try:
                    dt = datetime.strptime(value, "%Y-%m-%d")
                    week_dates.append(dt)
                except:
                    pass
        weeks[col] = sorted(week_dates)

    return weeks

def detect_current_term():
    """تحديد الفصل الصحيح حسب تاريخ اليوم مقارنة بأول تاريخ داخل كل شيت."""
    wb = openpyxl.load_workbook(EXCEL_PATH)

    terms = ["Term1", "Term2"]
    today = datetime.today()

    first_dates = {}

    for term in terms:
        ws = wb[term]
        # أول أسبوع = العمود 1
        # أول يوم = row 2
        value = ws.cell(row=2, column=1).value

        if value and not value.startswith("—"):
            try:
                dt = datetime.strptime(value, "%Y-%m-%d")
                first_dates[term] = dt
            except:
                pass

    # الآن نقارن بين التواريخ
    if len(first_dates) == 2:
        if today < first_dates["Term2"]:
            return "Term1"
        else:
            return "Term2"

    # fallback إذا لم تكتمل التواريخ
    return "Term1"

def detect_current_week(sheet_name="Term1"):
    """إرجاع رقم الأسبوع الحالي، وإن كان بين أسبوعين يرجّع السابق."""
    weeks = load_weeks_from_excel(sheet_name)
    today = datetime.today()

    last_week_with_dates = None

    for week_number, dates in weeks.items():
        if not dates:
            continue

        start = dates[0]
        end = dates[-1]

        # اليوم داخل هذا الأسبوع？
        if start <= today <= end:
            return week_number

        # اليوم بعد الأسبوع → مرشح كأسبوع سابق
        if end < today:
            last_week_with_dates = week_number

    return last_week_with_dates

# =========================
# تهيئة الفهارس (تشغل بالخلفية)
# =========================
INDEXES = {
    "majors_plan": {},
    "schedule": {},
    "advisor": None,
    "remaining": {},
    "gpa": {},
    "majors": {},
    "ids": {},
    "certificates": {},  # ← أضف هذا السطر
    "permission": {},
    "aramco_training": {},
}

# =========================
# فهرسة PDF (مع تقدم لحظي)
# =========================
def build_index(pdf_path, index_path="schedule_index.json"):
    """فهرسة ملف الجدول (Schedule) لاستخراج مواقع المتدربين حسب أرقامهم."""
    _set_status(indexing=True, current_file=os.path.basename(pdf_path), index_progress=0.0)
    try:
        cached = _load_index_if_current(pdf_path, index_path, f"✅ فهرس {pdf_path} جاهز مسبقًا.")
        if cached is not None:
            return cached

        if not os.path.exists(pdf_path):
            print(f"⚠️ الملف {pdf_path} غير موجود.", flush=True)
            return {}

        print(f"⏳ فهرسة الملف: {pdf_path}", flush=True)
        reader = PdfReader(pdf_path)
        total_pages = len(reader.pages)
        index = {}
        start_time = time.time()

        for i, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            for m in re.findall(r"\b44\d{7}\b", text):
                if m not in index:
                    index[m] = i - 1
            percent = (i / total_pages) * 100
            _set_status(index_progress=percent)
            if i % 100 == 0 or i == total_pages:
                print(f"📄 فهرسة schedule: الصفحة {i}/{total_pages} ({percent:.1f}%)", flush=True)

        with open(index_path, "w", encoding="utf-8") as f:
            json.dump(index, f, ensure_ascii=False)
        _write_index_meta(pdf_path, index_path)

        elapsed = time.time() - start_time
        print(f"✅ تم فهرسة {len(index)} متدرب من {pdf_path} خلال {elapsed:.1f} ثانية.", flush=True)
        return index
    except Exception as e:
        print("❌ خطأ أثناء فهرسة الجدول:", e, flush=True)
        import traceback; traceback.print_exc()
        return {}
    finally:
        _set_status(indexing=False, current_file="", index_progress=0.0)

def build_remaining_index(pdf_path, index_path="remaining_index.json"):
    _set_status(indexing=True, current_file=os.path.basename(pdf_path), index_progress=0.0)
    try:
        cached = _load_index_if_current(pdf_path, index_path, f"✅ فهرس {pdf_path} جاهز مسبقًا.")
        if cached is not None:
            return cached

        print(f"⏳ فهرسة (remaining) الملف: {pdf_path}", flush=True)
        reader = PdfReader(pdf_path)
        total_pages = len(reader.pages)
        index = {}
        start_time = time.time()

        for i, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            # ✅ اجمع أرقام المتدربين في هذه الصفحة كـ set لمنع تكرار نفس الصفحة
            # (لو ظهر الرقم مرتين/أكثر داخل نفس الصفحة نضيف رقم الصفحة مرة واحدة فقط)
            page_ids = set(re.findall(r"\b44\d{7}\b", text))
            for match in page_ids:
                index.setdefault(match, []).append(i - 1)
            percent = (i / total_pages) * 100
            _set_status(index_progress=percent)
            if i % 100 == 0 or i == total_pages:
                print(f"فهرسة remaining: الصفحة {i}/{total_pages} ({percent:.1f}%)", flush=True)

        with open(index_path, "w", encoding="utf-8") as f:
            json.dump(index, f, ensure_ascii=False)
        _write_index_meta(pdf_path, index_path)

        elapsed = time.time() - start_time
        print(f"✅ تم بناء فهرس remaining ({len(index)} متدرب) خلال {elapsed:.1f} ثانية.", flush=True)
        return index
    except Exception as e:
        print("❌ خطأ أثناء فهرسة remaining:", e, flush=True)
        import traceback; traceback.print_exc()
        return {}
    finally:
        _set_status(indexing=False, current_file="", index_progress=0.0)


# ==================================================
# ✉️ فهرسة خطابات التمكين
# ==================================================
def build_permission_index(pdf_path, index_path="permission_index.json"):
    """
    فهرسة ملف خطابات التمكين بناءً على رقم المتدرب (44xxxxxxx).
    يُنشئ خريطة: {student_id: [page_idx, ...]}
    مع ملف meta لإعادة البناء فقط عند تحديث الـ PDF.
    """
    _set_status(indexing=True, current_file=os.path.basename(pdf_path), index_progress=0.0)
    try:
        cached = _load_index_if_current(pdf_path, index_path, f"✅ فهرس {pdf_path} (permission) جاهز مسبقًا.")
        if cached is not None:
            return cached

        if not os.path.exists(pdf_path):
            print(f"⚠️ ملف خطابات التمكين غير موجود: {pdf_path}", flush=True)
            return {}

        print(f"⏳ فهرسة (permission) الملف: {pdf_path}", flush=True)
        reader = PdfReader(pdf_path)
        total_pages = len(reader.pages)
        index = {}
        start_time = time.time()

        for i, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""

            # ✅ اجمع أرقام المتدربين في هذه الصفحة كـ set لمنع تكرار نفس الصفحة
            # (لو ظهر الرقم مرتين/أكثر داخل نفس الصفحة نضيف رقم الصفحة مرة واحدة فقط)
            matches = set(re.findall(r"\b44\d{7}\b", text))
            for match in matches:
                index.setdefault(match, []).append(i - 1)

            percent = (i / total_pages) * 100 if total_pages else 100.0
            _set_status(index_progress=percent)
            if i % 10 == 0 or i == total_pages:
                print(f"فهرسة permission: الصفحة {i}/{total_pages} ({percent:.1f}%)", flush=True)

        with open(index_path, "w", encoding="utf-8") as f:
            json.dump(index, f, ensure_ascii=False)
        _write_index_meta(pdf_path, index_path)

        elapsed = time.time() - start_time
        print(f"✅ تم بناء فهرس permission ({len(index)} متدرب) خلال {elapsed:.1f} ثانية.", flush=True)
        return index

    except Exception as e:
        print("❌ خطأ أثناء فهرسة permission:", e, flush=True)
        import traceback; traceback.print_exc()
        return {}
    finally:
        _set_status(indexing=False, current_file="", index_progress=0.0)


# ==================================================
# 🏢 فهرسة خطابات تدريب أرامكو
# ==================================================
def build_aramco_training_index(pdf_path, index_path="aramco_training_index.json"):
    """
    فهرسة ملف خطابات تدريب أرامكو بناءً على رقم المتدرب (44xxxxxxx).
    النتيجة: {student_id: [page_idx, ...]}.
    """
    _set_status(indexing=True, current_file=os.path.basename(pdf_path), index_progress=0.0)
    try:
        cached = _load_index_if_current(pdf_path, index_path, f"✅ فهرس {pdf_path} (aramco training) جاهز مسبقًا.")
        if cached is not None:
            return cached

        if not os.path.exists(pdf_path):
            print(f"⚠️ ملف خطابات تدريب أرامكو غير موجود: {pdf_path}", flush=True)
            return {}

        print(f"⏳ فهرسة (aramco training) الملف: {pdf_path}", flush=True)
        reader = PdfReader(pdf_path)
        total_pages = len(reader.pages)
        index = {}
        start_time = time.time()

        for i, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            matches = set(re.findall(r"\b44\d{7}\b", text))
            for match in matches:
                index.setdefault(match, []).append(i - 1)

            percent = (i / total_pages) * 100 if total_pages else 100.0
            _set_status(index_progress=percent)
            if i % 10 == 0 or i == total_pages:
                print(f"فهرسة aramco training: الصفحة {i}/{total_pages} ({percent:.1f}%)", flush=True)

        with open(index_path, "w", encoding="utf-8") as f:
            json.dump(index, f, ensure_ascii=False)
        _write_index_meta(pdf_path, index_path)

        elapsed = time.time() - start_time
        print(f"✅ تم بناء فهرس aramco training ({len(index)} متدرب) خلال {elapsed:.1f} ثانية.", flush=True)
        return index

    except Exception as e:
        print("❌ خطأ أثناء فهرسة aramco training:", e, flush=True)
        import traceback; traceback.print_exc()
        return {}
    finally:
        _set_status(indexing=False, current_file="", index_progress=0.0)

# ==================================================
# 📘 فهرسة الشهادات   
# ==================================================
def build_certificates_index(pdf_path, index_path="certificates_index.json"):
    """
    فهرسة الشهادات باستخراج رقم الهوية الإنجليزي مباشرة من نص PDF بدون OCR.
    """
    index = {}

    if not os.path.exists(pdf_path):
        print(f"❌ ملف الشهادات غير موجود: {pdf_path}", flush=True)
        return index

    try:
        cached = _load_index_if_current(pdf_path, index_path, "✅ فهرس الشهادات جاهز مسبقًا.")
        if cached is not None:
            return cached

        print("🔍 بدء فهرسة الشهادات (بدون OCR)...", flush=True)
        reader = PdfReader(pdf_path)
        total = len(reader.pages)

        for i, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""

            # 🔴 Regex لرقم هوية سعودي 10 أرقام يبدأ بـ 1
            matches = re.findall(r"\b1\d{9}\b", text)

            if matches:
                for nid in matches:
                    index.setdefault(nid, []).append(i - 1)
            if i % 10 == 0 or i == total:
                percent = (i / total) * 100 if total else 100.0
                print(f"فهرسة certificates: الصفحة {i}/{total} ({percent:.1f}%)", flush=True)

        # حفظ الفهرس
        with open(index_path, "w", encoding="utf-8") as f:
            json.dump(index, f, ensure_ascii=False, indent=2)
        _write_index_meta(pdf_path, index_path)

        print(f"✅ تم بناء فهرس الشهادات بنجاح – عدد الهويات: {len(index)}")
        return index

    except Exception as e:
        print("❌ خطأ أثناء فهرسة الشهادات:", e, flush=True)
        return {}

def load_ids_from_csv(csv_path: str):
    """
    🔹 تحميل بيانات المتدربين من ملف CSV يحتوي على الأعمدة:
    الفصل التدريبي,"الوحدة التدريبية","المرحلة","القسم","البرنامج",
    "رقم المتدرب","اسم المتدرب","المعدل التراكمي","السجل المدني","الجنس","الجنسية","رقم الجوال"
    🔸 النتيجة: {"رقم المتدرب": {"nid": "السجل المدني", "name": "اسم المتدرب"}}
    """
    index = {}
    if not os.path.exists(csv_path):
        print(f"⚠️ ملف CSV غير موجود: {csv_path}", flush=True)
        return index

    try:
        df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str, quotechar='"')
        for _, row in df.iterrows():
            sid = str(row.get("رقم المتدرب", "")).strip()
            nid = str(row.get("السجل المدني", "")).strip()
            nid = normalize_arabic_text_v2(nid)
            name = str(row.get("اسم المتدرب", "")).strip()
            if re.fullmatch(r"44\d{7}", sid) and re.fullmatch(r"1\d{9}", nid):
                index[sid] = {"nid": nid, "name": name}
        print(f"✅ تم تحميل بيانات {len(index)} متدرب من CSV بنجاح.", flush=True)
    except Exception as e:
        print(f"❌ خطأ أثناء قراءة CSV: {e}", flush=True)
        import traceback; traceback.print_exc()

    return index


def build_majors_index(pdf_path, index_path="majors_index.json"):
    try:
        cached = _load_index_if_current(pdf_path, index_path, "✅ فهرس التخصصات جاهز مسبقًا.")
        if cached is not None:
            return cached

        print(f"🔍 بناء فهرس التخصصات {pdf_path} ...", flush=True)
        reader = PdfReader(pdf_path)
        index = {}
        total_pages = len(reader.pages)

        for i, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            student_ids = re.findall(r"\b44\d{7}\b", text)
            if student_ids:
                for sid in student_ids:
                    index[sid] = text
            if i % 100 == 0 or i == total_pages:
                percent = (i / total_pages) * 100
                print(f"📄 فهرسة الصفحة {i}/{total_pages} ({percent:.1f}%)", flush=True)

        with open(index_path, "w", encoding="utf-8") as f:
            json.dump(index, f, ensure_ascii=False)
        _write_index_meta(pdf_path, index_path)

        print(f"✅ تم بناء فهرس التخصصات ({len(index)} متدرب).", flush=True)
        return index
    except Exception as e:
        print("❌ خطأ أثناء فهرسة التخصصات:", e, flush=True)
        import traceback; traceback.print_exc()
        return {}


def initialize_indexes():
    print("🚀 بدء تشغيل النظام وفهرسة الملفات بالخلفية...", flush=True)
    INDEXES_READY.clear()
    _set_status(indexes_ready=False)
    try:
        print("\n📂 فهرسة SCHEDULE ...", flush=True)
        INDEXES["schedule"] = build_index(FILES["schedule"])

        print("\n📂 فهرسة REMAINING ...", flush=True)
        INDEXES["remaining"] = build_remaining_index(FILES["remaining"])

        INDEXES["gpa"] = {}

        print("\n📂 فهرسة IDs ...", flush=True)
        INDEXES["ids"] = load_ids_from_csv("IDs.csv")

        print("\n📂 فهرسة MAJORS ...", flush=True)
        INDEXES["majors"] = build_majors_index(FILES["majors"])

        # ✅ فهرس سريع: رقم المتدرب -> ملف الخطة (window=1)
        print("\n📂 فهرسة MAJORS PLAN ...", flush=True)
        INDEXES["majors_plan"] = build_majors_plan_index(FILES["majors"])

        print("\n📂 فهرسة CERTIFICATES ...", flush=True)
        INDEXES["certificates"] = build_certificates_index(FILES["certificates"])
        print("\n📂 فهرسة PERMISSION LETTERS ...", flush=True)
        INDEXES["permission"] = build_permission_index(FILES["permission"])

        print("\n📂 فهرسة ARAMCO TRAINING LETTERS ...", flush=True)
        INDEXES["aramco_training"] = build_aramco_training_index(FILES["aramco_training"])

        
        INDEXES["advisor"] = None
        INDEXES_READY.set()
        _set_status(indexes_ready=True)
        print("\n----------------------------", flush=True)
        print("✅ جميع الفهارس جاهزة بنجاح.", flush=True)

    except Exception as e:
        INDEXES_READY.clear()
        _set_status(indexes_ready=False)
        print("❌ خطأ أثناء التهيئة:", e, flush=True)
        import traceback
        traceback.print_exc()

# =========================
# ضغط PDF
# =========================
def _gs_binary():
    # استخدم gswin64c على ويندوز، و gs على أنظمة أخرى
    return "gswin64c" if os.name == "nt" else "gs"

def compress_pdf_with_ghostscript(input_file: str, output_file: str, max_size_mb: float = 3.0):
    """ضغط PDF بواسطة Ghostscript مع خطة بديلة."""
    print(f"⏳ ضغط الملف {input_file} ...", flush=True)
    try:
        command = [
            _gs_binary(), "-sDEVICE=pdfwrite", "-dCompatibilityLevel=1.4",
            "-dPDFSETTINGS=/ebook", "-dNOPAUSE", "-dQUIET", "-dBATCH",
            f"-sOutputFile={output_file}", input_file
        ]
        subprocess.run(command, check=True)
        size_mb = os.path.getsize(output_file) / (1024 * 1024)
        print(f"✅ تم ضغط الملف ({size_mb:.2f} MB) باستخدام إعداد /ebook", flush=True)
        return True
    except Exception as e:
        print(f"⚠️ فشل الضغط الأول ({e})، تجربة إعداد /screen...", flush=True)
        try:
            command = [
                _gs_binary(), "-sDEVICE=pdfwrite", "-dCompatibilityLevel=1.4",
                "-dPDFSETTINGS=/screen", "-dNOPAUSE", "-dQUIET", "-dBATCH",
                f"-sOutputFile={output_file}", input_file
            ]
            subprocess.run(command, check=True)
            size_mb = os.path.getsize(output_file) / (1024 * 1024)
            print(f"✅ تم ضغط الملف ({size_mb:.2f} MB) باستخدام إعداد /screen", flush=True)
            return True
        except Exception as e2:
            print(f"❌ فشل الضغط تمامًا ({e2})، سيتم استخدام النسخة الأصلية.", flush=True)
            return False

# =========================
# الخدمات
# =========================
async def safe_delete_message(message):
    try:
        await message.delete()
    except Exception:
        pass

class UserFacingError(Exception):
    pass

def _make_temp_dir(service: str) -> str:
    os.makedirs(TEMP_ROOT, exist_ok=True)
    tmpdir = os.path.join(TEMP_ROOT, f"tvtc_{service}_{uuid.uuid4().hex}")
    os.makedirs(tmpdir, exist_ok=False)
    return tmpdir

def _write_pdf_subset(reader: PdfReader, pages, output_file: str):
    writer = PdfWriter()
    for page_index in pages:
        writer.add_page(reader.pages[page_index])
    with open(output_file, "wb") as f:
        writer.write(f)

def _compress_or_original(output_file: str, compressed_file: str) -> str:
    success = compress_pdf_with_ghostscript(output_file, compressed_file)
    if not success:
        print("⚠️ فشل الضغط، سيتم إرسال النسخة الأصلية.", flush=True)
        return output_file
    return compressed_file

def _prepare_pdf_document(service: str, student_id: str):
    pdf_path = FILES.get(service)
    index = INDEXES.get(service) or {}
    if not pdf_path or not os.path.exists(pdf_path):
        raise UserFacingError("❌ الملف المطلوب غير متاح حالياً.")

    tmpdir = _make_temp_dir(service)
    try:
        reader = PdfReader(pdf_path)

        if service == "certificates":
            ids_map = INDEXES.get("ids") or {}
            student_info = ids_map.get(student_id, {})
            nid_clean = normalize_arabic_text_v2(str(student_info.get("nid")))
            pages = sorted(set(index.get(nid_clean, [])))
            if not nid_clean or not pages:
                raise UserFacingError("⚠️ لم يتم العثور على شهادات لهذا المتدرب.")

            output_file = os.path.join(tmpdir, "certificates.pdf")
            compressed_file = os.path.join(tmpdir, "compressed_certificates.pdf")
            _write_pdf_subset(reader, pages, output_file)
            document_path = _compress_or_original(output_file, compressed_file)
            return {
                "path": document_path,
                "filename": f"certificates_{student_id}.pdf",
                "caption": f"📜 شهادات البرامج الخاصة بالمتدرب رقم {student_id}",
                "tmpdir": tmpdir,
            }

        if service in ("permission", "aramco_training"):
            service_config = {
                "permission": {
                    "missing": "⚠️ لا يوجد خطاب تمكين مرتبط بهذا الرقم التدريبي.",
                    "output": "permission.pdf",
                    "compressed": "compressed_permission.pdf",
                    "filename": f"permission_{student_id}.pdf",
                    "caption": f"✉️ خطاب التمكين للمتدرب رقم {student_id}",
                },
                "aramco_training": {
                    "missing": "⚠️ لا يوجد خطاب تدريب أرامكو مرتبط بهذا الرقم التدريبي.",
                    "output": "aramco_training.pdf",
                    "compressed": "compressed_aramco_training.pdf",
                    "filename": f"aramco_training_{student_id}.pdf",
                    "caption": f"🏢 خطاب تدريب أرامكو للمتدرب رقم {student_id}",
                },
            }[service]

            pages = sorted(set(index.get(student_id, [])))
            if not pages:
                raise UserFacingError(service_config["missing"])

            output_file = os.path.join(tmpdir, service_config["output"])
            compressed_file = os.path.join(tmpdir, service_config["compressed"])
            _write_pdf_subset(reader, pages, output_file)
            document_path = _compress_or_original(output_file, compressed_file)
            return {
                "path": document_path,
                "filename": service_config["filename"],
                "caption": service_config["caption"],
                "tmpdir": tmpdir,
            }

        if service == "remaining":
            pages = sorted(set(index.get(student_id, [])))
            if not pages:
                raise UserFacingError(f"❌ لم يتم العثور على مقررات المتدرب {student_id}.")
        else:
            if student_id not in index:
                raise UserFacingError("❌ لم يتم العثور على بياناتك.")

            start = index[student_id]
            sorted_students = sorted(index.items(), key=lambda x: x[1])
            end = len(reader.pages)
            for _, page_idx in sorted_students:
                if page_idx > start:
                    end = page_idx
                    break
            pages = range(start, end)

        output_file = os.path.join(tmpdir, f"{service}.pdf")
        compressed_file = os.path.join(tmpdir, f"compressed_{service}.pdf")
        _write_pdf_subset(reader, pages, output_file)
        document_path = _compress_or_original(output_file, compressed_file)

        captions = {
            "schedule": f"📄 جدول المتدرب رقم {student_id}",
            "remaining": f"📚 المقررات المتبقية للمتدرب رقم {student_id}",
        }
        return {
            "path": document_path,
            "filename": f"{service}_{student_id}.pdf",
            "caption": captions.get(service, f"📄 ملف {service} للمتدرب {student_id}"),
            "tmpdir": tmpdir,
        }

    except Exception:
        shutil.rmtree(tmpdir, ignore_errors=True)
        raise

async def _reply_document_with_retry(message, document_path: str, filename: str, caption: str):
    for attempt in range(2):
        try:
            with open(document_path, "rb") as f:
                await message.reply_document(
                    f,
                    filename=filename,
                    caption=caption
                )
            return
        except (NetworkError, TimedOut):
            if attempt == 1:
                raise
            await asyncio.sleep(2)

def _lookup_advisor(student_id: str):
    csv_path = FILES.get("advisor")
    df = pd.read_csv(csv_path, encoding='utf-8', dtype=str)

    advisor_name = None
    mask = df.apply(lambda row: row.astype(str).str.contains(student_id, regex=False, na=False).any(), axis=1)
    matched_rows = df[mask]
    if not matched_rows.empty:
        for _, row in matched_rows.iterrows():
            text = " ".join(row.dropna().astype(str))
            match = re.search(r"00\d{5,7}\s*([^\d\n\r]+)", text)
            if match:
                advisor_name = match.group(1).strip()
                advisor_name = re.sub(r"مرشد أكاديمي", "", advisor_name)
                advisor_name = advisor_name.replace(",", "").replace('"', "").strip()
                break
    return advisor_name

def _lookup_gpa(student_id: str):
    reader = PdfReader(FILES.get("gpa"))
    for page in reader.pages:
        text = page.extract_text() or ""
        for line in text.splitlines():
            if student_id in line:
                match = re.search(r"\b\d\.\d{2}\b", line)
                if match:
                    return match.group(0)
    return None

async def send_advisor(update, context, student_id):
    csv_path = FILES.get("advisor")
    if not os.path.exists(csv_path):
        await update.message.reply_text("❌ ملف المرشد غير متاح حالياً.")
        return
    sent_msg = await update.message.reply_text("👨‍🏫 جاري البحث عن مرشدك التدريبي...")
    try:
        advisor_name = await asyncio.to_thread(_lookup_advisor, student_id)
    except Exception as e:
        await safe_delete_message(sent_msg)
        await update.message.reply_text("❌ خطأ في قراءة ملف المرشدين.")
        import traceback
        traceback.print_exc()
        return

    await safe_delete_message(sent_msg)
    if advisor_name:
        await update.message.reply_text(f"👨‍🏫 مرشدك التدريبي هو:\nأ. {advisor_name}")
    else:
        await update.message.reply_text("⚠️ لم يتم العثور على اسم المرشد.")

async def send_gpa(update, context, student_id):
    pdf_path = FILES.get("gpa")
    if not os.path.exists(pdf_path):
        await update.message.reply_text("❌ ملف المعدل غير متاح حالياً.")
        return
    sent_msg = await update.message.reply_text("🎓 جاري البحث عن معدلك...")
    try:
        gpa_value = await asyncio.to_thread(_lookup_gpa, student_id)
    except Exception as e:
        await safe_delete_message(sent_msg)
        await update.message.reply_text("❌ خطأ في قراءة ملف المعدل.")
        import traceback
        traceback.print_exc()
        return

    await safe_delete_message(sent_msg)
    if gpa_value:
        await update.message.reply_text(f"🎓 معدلك هو: {gpa_value}")
    else:
        await update.message.reply_text("⚠️ لم يتم العثور على المعدل.")

# خرائط العبارات إلى ملفات الخطط + كابتشنات
MAJOR_PHRASES_TO_PLAN = {
    "قيرتتلارصاتت لاااتتلاا يرت": "VocationalSafetyAndHealth.pdf",
    "قيراتلااا لقلبتلا رارضترا": "LabsPlan.pdf",
    "لاارلقمتلالاقرل": "HRplan.pdf",
    "فاارتتلالرقت": "EPplan.pdf",
    "قارتترماتتلةينرت": "FoodSafetyPlan.pdf",
}

PLAN_CAPTIONS = {
    "HRplan.pdf": "💼 الخطة التفصيلية لتخصص الموارد البشرية",
    "EPplan.pdf": "🌿 الخطة التفصيلية لتخصص حماية البيئة",
    "FoodSafetyPlan.pdf": "🍽️ الخطة التفصيلية لتخصص سلامة الأغذية",
    "LabsPlan.pdf": "🧪 الخطة التفصيلية لتخصص المختبرات الكيميائية",
    "VocationalSafetyAndHealth.pdf": "🦺 الخطة التفصيلية لتخصص السلامة والصحة المهنية",
}

def _normalize_spaces(s: str) -> str:
    return " ".join((s or "").split())

async def send_detailed_plan(update, context, student_id):
    # لو جاء الطلب من زر inline لازم نجاوب الـ callback
    query = update.callback_query
    if query:
        await query.answer()

    # effective_message تشتغل سواء كان update رسالة عادية أو callback
    msg = update.effective_message

    plan_index = INDEXES.get("majors_plan") or {}
    plan_file_to_send = plan_index.get(str(student_id).strip())

    if not plan_file_to_send:
        await msg.reply_text("⚠️ لم يتم العثور على التخصص المناسب.")
        return

    if not os.path.exists(plan_file_to_send):
        await msg.reply_text(f"⚠️ تم تحديد الخطة ({plan_file_to_send}) لكن الملف غير موجود في السيرفر.")
        return

    caption = PLAN_CAPTIONS.get(plan_file_to_send, "📑 خطتك التفصيلية")
    try:
        with open(plan_file_to_send, "rb") as f:
            await msg.reply_document(
                document=f,
                filename=os.path.basename(plan_file_to_send),
                caption=caption
            )
    except Exception:
        await msg.reply_text("⚠️ تعذر إرسال الملف. تحقق من صلاحيات القراءة أو حجم الملف.")


async def send_pdf(update: Update, context: ContextTypes.DEFAULT_TYPE, service: str):
    student_id = context.user_data.get("student_id")
    if not student_id:
        await update.message.reply_text("⚠️ الرجاء إدخال رقمك التدريبي أولاً.")
        return

    if not INDEXES_READY.is_set():
        await _reply_indexes_not_ready(update)
        return

    if service == "advisor":
        await send_advisor(update, context, student_id)
        return
    if service == "gpa":
        await send_gpa(update, context, student_id)
        return
    if service == "detailed_plan":
        await send_detailed_plan(update, context, student_id)
        return

    messages = {
        "schedule": "📄 جاري تجهيز جدولك...",
        "remaining": "📚 جاري حصر مقرراتك المتبقية...",
        "certificates": "📜 جاري تجهيز شهاداتك...",
        "permission": "✉️ جاري تجهيز خطاب التمكين...",
        "aramco_training": "🏢 جاري تجهيز خطاب تدريب أرامكو...",
    }
    sent_msg = await update.message.reply_text(messages.get(service, "⏳ جاري تجهيز الملف..."))

    prepared = None

    try:
        prepared = await asyncio.to_thread(_prepare_pdf_document, service, student_id)
        await _reply_document_with_retry(
            update.message,
            prepared["path"],
            prepared["filename"],
            prepared["caption"]
        )

    except UserFacingError as e:
        await update.message.reply_text(str(e))
    except (NetworkError, TimedOut):
        await update.message.reply_text("⚠️ تعذر إرسال الملف بسبب مهلة الاتصال. حاول مرة أخرى.")
    except Exception as e:
        await update.message.reply_text("❌ حدث خطأ أثناء تجهيز الملف. حاول مرة أخرى لاحقاً.")
        import traceback
        traceback.print_exc()

    finally:
        await safe_delete_message(sent_msg)
        if prepared and prepared.get("tmpdir"):
            shutil.rmtree(prepared["tmpdir"], ignore_errors=True)

# =========================
# دالة مساعدة لبناء لوحة الأزرار
# =========================
def build_main_keyboard(student_id: str):
    """بناء لوحة الخدمات بناءً على حالة المتدرب (هل له مقررات أو شهادات)."""
    has_remaining = student_id in INDEXES.get("remaining", {})
    has_permission = student_id in INDEXES.get("permission", {})
    has_aramco_training = student_id in INDEXES.get("aramco_training", {})

    # نحصل على رقم الهوية من فهرس IDs
    ids_map = INDEXES.get("ids", {})
    rec = ids_map.get(student_id, {})
    nid = rec.get("nid", "")

    # هل يوجد شهادة بناء على رقم الهوية؟
    has_certificate = False

    # ✅ تطبيع رقم الهوية من CSV
    nid_clean = normalize_arabic_text_v2(str(nid)).strip()

    # ✅ المقارنة مع فهرس الشهادات
    if nid_clean and nid_clean in INDEXES.get("certificates", {}):
        has_certificate = True

    keyboard = [
        [KeyboardButton("📄 جدولي")],
        [KeyboardButton("👨‍🏫 مرشدي التدريبي"), KeyboardButton("🎓 معدلي")],
        [KeyboardButton("📑 خطتي التفصيلية")],
        [KeyboardButton("📅 الأسبوع الحالي")],
        [KeyboardButton("📤 تسجيل الخروج")]
    ]

    # نضيف المقررات إن وجدت
    if has_remaining:
        keyboard[0].append(KeyboardButton("📚 مقرراتي المتبقية"))

    insert_at = 1

    # نضيف زر الشهادات فقط إذا له شهادة
    if has_certificate:
        keyboard.insert(insert_at, [KeyboardButton("📜 شهادات البرامج المساندة")])
        insert_at += 1

    # نضيف زر خطابات التمكين فقط إذا له خطاب
    if has_permission:
        keyboard.insert(insert_at, [KeyboardButton("✉️ خطابات التمكين")])
        insert_at += 1

    # نضيف زر خطاب تدريب أرامكو فقط إذا له خطاب
    if has_aramco_training:
        keyboard.insert(insert_at, [KeyboardButton("🏢 خطاب تدريب أرامكو")])

    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

# =========================
# معالجات الرسائل
# =========================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # 🧹 1) إلغاء أي مهام عدّ تنازلي
    task = context.user_data.pop("logout_task", None)
    if task:
        try:
            task.cancel()
        except:
            pass

    # 🧹 2) مسح الحالة بالكامل
    context.user_data.clear()

    # 🧹 3) إزالة لوحة الأزرار والرجوع للبداية
    await update.message.reply_text(
        "🔄 تم البدء من جديد.\n\n"
        "👋 أرسل رقمك التدريبي (يبدأ بـ 44 ويتكون من 9 أرقام) للحصول على خدماتك.",
        reply_markup=ReplyKeyboardRemove()
    )

# =========================
# دالة العد التنازلي بعد تسجيل الخروج
# =========================
async def countdown_message(msg, chat_id, context: ContextTypes.DEFAULT_TYPE):
    try:
        clock_emojis = ["🕐","🕑","🕒","🕓","🕔","🕕","🕖","🕗","🕘","🕙","🕚","🕛"]

        for remaining in range(59, 0, -1):
            await asyncio.sleep(1)
            clock = clock_emojis[remaining % len(clock_emojis)]

            try:
                await msg.edit_text(
                    f"✅ تم تسجيل خروجك بنجاح.\n\n"
                    f"يمكنك إدخال رقم تدريبي جديد أو إعادة تسجيل الدخول قبل {remaining} ثانية {clock}",
                    reply_markup=InlineKeyboardMarkup([
                        [InlineKeyboardButton("🔁 إعادة تسجيل الدخول", callback_data='relogin')]
                    ])
                )
            except:
                break

        try:
            await msg.delete()
        except:
            pass

        await context.bot.send_message(
            chat_id=chat_id,
            text=(
                "👋 مرحباً!\n"
                "أرسل رقمك التدريبي (يبدأ بـ 44 ويتكون من 9 أرقام) للحصول على خدماتك."
            ),
            reply_markup=ReplyKeyboardRemove()
        )

    except asyncio.CancelledError:
        pass
    except Exception as e:
        print("⚠️ خطأ في العد التنازلي:", e, flush=True)

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    txt = (update.message.text or "").strip()
    student_id = convert_arabic_to_english(txt)
    _log_user_input(txt, context)

    # 🚨 إذا كان المستخدم في فترة العد التنازلي وقام بإرسال أي رسالة → ألغِ العد فوراً
    if "logout_task" in context.user_data:
        task = context.user_data.pop("logout_task", None)
        msg_id = context.user_data.pop("logout_message_id", None)

        # إلغاء مهمة العد التنازلي
        if task:
            try:
                task.cancel()
            except:
                pass

        # حذف رسالة العد
        if msg_id:
            try:
                await context.bot.delete_message(
                    chat_id=update.effective_chat.id,
                    message_id=msg_id
                )
            except:
                pass

        # إعادة المستخدم لشاشة البداية
        await update.message.reply_text(
            "👋 مرحباً!\n"
            "أرسل رقمك التدريبي (يبدأ بـ 44 ويتكون من 9 أرقام) للحصول على خدماتك.",
            reply_markup=ReplyKeyboardRemove()
        )
        return

    if await _expire_session_if_needed(update, context):
        return

    # =============================
    # تسجيل الخروج
    # =============================
    if txt.strip() == "📤 تسجيل الخروج":
        last_id = context.user_data.get("student_id")
        context.user_data.clear()

        if last_id:
            context.user_data["last_student_id"] = last_id

        await update.message.reply_text(
            "جارٍ تسجيل الخروج...",
            reply_markup=ReplyKeyboardRemove()
        )
        await asyncio.sleep(0.3)

        sent_msg = await update.message.reply_text(
            "✅ تم تسجيل خروجك بنجاح.\n\n"
            "يمكنك إدخال رقم تدريبي جديد أو إعادة تسجيل الدخول خلال 60 ثانية 🕐",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔁 إعادة تسجيل الدخول", callback_data="relogin")]
            ])
        )

        # تشغيل العدّ التنازلي بالخلفية
        task = asyncio.create_task(countdown_message(sent_msg, update.effective_chat.id, context))
        context.user_data["logout_task"] = task
        context.user_data["logout_message_id"] = sent_msg.message_id
        return

    if not await _ensure_indexes_ready(update):
        return

        # ========= مرحلة التحقق على خطوتين =========
    # 1️⃣ المستخدم أدخل رقم متدرب صالح 44xxxxxxx
    if re.match(r"^44\d{7}$", student_id):
        if "student_id" in context.user_data:
            await update.message.reply_text("⚠️ يرجى تسجيل الخروج أولًا قبل إدخال رقم جديد.")
            return

        context.user_data["pending_student_id"] = student_id
        await update.message.reply_text("🪪 أدخل رقم الهوية الوطنية (10 أرقام):")
        return

    # 2️⃣ المستخدم في وضع انتظار الهوية وأرسل 10 أرقام
    if "pending_student_id" in context.user_data and re.match(r"^[0-9٠-٩]{10}$", txt):
        entered_nid = normalize_digits(txt)

        if not is_valid_nid(entered_nid):
            await update.message.reply_text("⚠️ أدخل رقم الهوية الوطنية الصحيح (10 أرقام يبدأ بـ 1).")
            return

        pending_id = context.user_data.get("pending_student_id")
        ids_map = INDEXES.get("ids") or {}
        rec = ids_map.get(pending_id)

        if not rec or "nid" not in rec:
            await update.message.reply_text("⚠️ لا توجد بيانات هوية مرتبطة بهذا الرقم التدريبي. تواصل مع الدعم.")
            return

        if normalize_digits(str(rec["nid"])) != entered_nid:
            await update.message.reply_text("❌ رقم الهوية غير مطابق لرقم المتدرب. حاول مرة أخرى.")
            return

        _mark_authenticated(context, pending_id)
        context.user_data.pop("pending_student_id", None)

        full_name = rec.get("name", "").strip()
        first_name = extract_first_name(full_name)

        keyboard = build_main_keyboard(pending_id)

        await update.message.reply_text(
            f"🎉 أهلاً وسهلاً {first_name}!\nالآن يمكنك الاستفادة من خدماتك:",
            reply_markup=keyboard
        )
        return

        # =============================
    # ⛔ منع استخدام الخدمات بدون تسجيل دخول
    # =============================
    protected_buttons = [
        "📜 شهادات البرامج المساندة",
        "📄 جدولي",
        "📚 مقرراتي المتبقية",
        "👨‍🏫 مرشدي التدريبي",
        "🎓 معدلي",
        "📑 خطتي التفصيلية",
        "📅 الأسبوع الحالي",
        "✉️ خطابات التمكين",
        "🏢 خطاب تدريب أرامكو"
    ]

    if txt in protected_buttons:
        # إذا لم يسجل دخول → نمنعه
        if "student_id" not in context.user_data:
            await update.message.reply_text("⚠️ الرجاء إدخال رقمك التدريبي أولاً.")
            return

        # زر الأسبوع الحالي
        if txt == "📅 الأسبوع الحالي":
            await current_week_handler(update, context)
            return

        # الخدمات الأخرى
        mapping = {
            "📄 جدولي": "schedule",
            "📚 مقرراتي المتبقية": "remaining",
            "👨‍🏫 مرشدي التدريبي": "advisor",
            "🎓 معدلي": "gpa",
            "📑 خطتي التفصيلية": "detailed_plan",
            "📜 شهادات البرامج المساندة": "certificates",
            "✉️ خطابات التمكين": "permission",
            "🏢 خطاب تدريب أرامكو": "aramco_training",
        }

        service = mapping.get(txt)
        if service:
            await send_pdf(update, context, service)
            return

    # أي رسالة أخرى غير مفهومة
    await update.message.reply_text(
        "⚠️ يرجى إدخال رقم تدريبي صحيح :\n"
        "(يبدأ بـ 44 ويتكون من 9 أرقام)"
    )

# تحويل رقم الأسبوع إلى نص عربي
WEEK_ARABIC = {
    1: "الأول",
    2: "الثاني",
    3: "الثالث",
    4: "الرابع",
    5: "الخامس",
    6: "السادس",
    7: "السابع",
    8: "الثامن",
    9: "التاسع",
    10: "العاشر",
    11: "الحادي عشر",
    12: "الثاني عشر",
    13: "الثالث عشر",
    14: "الرابع عشر",
    15: "الخامس عشر",
    16: "السادس عشر",
    17: "السابع عشر",
    18: "الثامن عشر",
    19: "التاسع عشر"
}

# =========================
# 🟦 هاندلر زر الأسبوع الحالي (النسخة الصحيحة النهائية)
# =========================
async def current_week_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    term = detect_current_term()
    week = detect_current_week(term)

    if week:
        week_name = WEEK_ARABIC.get(week, str(week))
        term_name = "الأول" if term == "Term1" else "الثاني"
        await update.message.reply_text(f"📅 نحن الآن في الأسبوع {week_name} – الفصل {term_name}")
    else:
        await update.message.reply_text("⚠️ لا يمكن تحديد الأسبوع حالياً.")

# =========================
# التشغيل الرئيسي
# =========================
def main():
    _set_status(running=True, telegram_connected=False)
    # شغّل الفهرسة بالخلفية
    threading.Thread(target=initialize_indexes, daemon=True).start()

    print("🚀 تشغيل البوت...", flush=True)
    app = (
        ApplicationBuilder()
        .token(BOT_TOKEN)
        .connect_timeout(30)
        .read_timeout(60)
        .write_timeout(120)
        .pool_timeout(30)
        .build()
    )

    # 🟢 معالجات الأوامر والرسائل
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    # زر الأسبوع الحالي
    app.add_handler(MessageHandler(filters.Regex("^الأسبوع الحالي$"), current_week_handler))


       # 🟢 تعريف الدالة التي تتعامل مع زر "اضغط هنا لإعادة تسجيل الدخول"
    async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        await query.answer()

        if query.data == "relogin":
            last_id = context.user_data.get("last_student_id")
            if not last_id:
                await query.edit_message_text("⚠️ لا يوجد رقم تدريبي سابق لإعادة تسجيل الدخول.")
                return
            if not INDEXES_READY.is_set():
                await query.edit_message_text("⏳ البوت يجهز بيانات الطلاب الآن. حاول بعد قليل.")
                return

            # ✅ حذف رسالة العدّ التنازلي مع الزر فورًا
            try:
                await query.delete_message()
            except Exception:
                pass

            # ✅ إعادة تخزين رقم المتدرب
            _mark_authenticated(context, last_id)

            # ✅ إرسال لوحة الخدمات فقط مع رسالة ترحيبية نظيفة
            keyboard = build_main_keyboard(last_id)

            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text=f"✅ تم تسجيل دخولك مجددًا بالرقم ({last_id}).\nاختر الخدمة:",
                reply_markup=keyboard
            )

    # 🟢 تفعيل معالج الأزرار بعد تعريف الدالة
    app.add_handler(CallbackQueryHandler(handle_callback))

    # =========================
    # تهيئة الاتصال بعد بدء التطبيق
    # =========================
    async def post_init(application):
        try:
            me = await application.bot.get_me()
            print(f" معلومات البوت: @{me.username} (id={me.id})", flush=True)
            _set_status(telegram_connected=True)
            await application.bot.set_my_commands([("start", "بدء البوت")])
        except Exception as e:
            print(f"⚠️ تعذر التأكد من اتصال تيليجرام: {e}", flush=True)
            _set_status(telegram_connected=False)

    app.post_init = post_init

    print("✅ البوت جاهز لاستقبال الطلبات الآن.", flush=True)

    # =========================
    # تشغيل البوت
    # =========================
    try:
        app.run_polling(allowed_updates=Update.ALL_TYPES)
    except KeyboardInterrupt:
        pass
    finally:
        _set_status(
            running=False,
            telegram_connected=False,
            indexing=False,
            current_file="",
            index_progress=0.0
        )
        print("👋 تم إيقاف البوت، يتم إنهاء جميع العمليات...", flush=True)
        try:
            import os as _os, signal as _signal
            _os.kill(_os.getpid(), _signal.SIGTERM)
        except Exception as e:
            print("⚠️ فشل إنهاء العملية:", e, flush=True)
        time.sleep(0.2)
        os._exit(0)


if __name__ == "__main__":
    main()
