import openpyxl
from datetime import datetime

EXCEL_PATH = "Calender.xlsx"

def load_weeks_from_excel(sheet_name):
    """تحميل الأسابيع من شيت محدد (Term1 أو Term2)."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[sheet_name]

    weeks = {}

    for col in range(1, ws.max_column + 1):
        week_dates = []
        for row in range(2, 7):  # 5 صفوف (الأحد → الخميس)
            value = ws.cell(row=row, column=col).value
            if value and not value.startswith("—"):
                try:
                    dt = datetime.strptime(value, "%Y-%m-%d")
                    week_dates.append(dt)
                except:
                    pass
        weeks[col] = sorted(week_dates)

    return weeks  # {1: [تواريخ], 2: [...], ...}


def detect_current_week(sheet_name="Term1"):
    """تحديد الأسبوع الحالي، وإذا كان بين أسبوعين يأخذ السابق."""
    
    weeks = load_weeks_from_excel(sheet_name)
    today = datetime.today()

    last_week_with_dates = None

    for week_number, dates in weeks.items():
        if not dates:
            continue
        
        start = dates[0]
        end = dates[-1]

        # 1) اليوم داخل أسبوع
        if start <= today <= end:
            return week_number

        # 2) الأسبوع انتهى قبل اليوم → مرشح للأسبوع السابق
        if end < today:
            last_week_with_dates = week_number

    # إن لم يكن اليوم داخل أسبوع → رجّع آخر أسبوع قبل اليوم
    return last_week_with_dates


# ------------------------- اختبار -------------------------

if __name__ == "__main__":
    term = "Term1"  # غيّرها إلى "Term2" للفصل الثاني
    
    week = detect_current_week(term)
    print(f"الأسبوع الحالي (أو السابق): Week {week}")
